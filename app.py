import streamlit as st
import pandas as pd
import json
import os
import textwrap
import io
from xhtml2pdf import pisa

st.set_page_config(page_title="GST Refund Automation Engine", layout="wide")

# Inject Custom CSS for Nav Bar Top Alignment
st.markdown("""
<style>
    .stApp {
        background-color: #FFFFFF;
    }
    .stApp > header {
        display: none !important;
    }
    .nav-bar {
        position: fixed;
        top: 0;
        left: 0;
        right: 0;
        width: 100vw;
        background: linear-gradient(90deg, #1E3A8A 0%, #3B82F6 100%);
        padding: 10px 30px !important;
        color: white;
        z-index: 999999;
        box-shadow: 0 4px 10px rgba(0,0,0,0.15);
    }
    .nav-bar h1 {
        color: white !important;
        margin: 0 !important;
        padding: 0 !important;
        font-size: 20px !important;
        font-family: 'Inter', sans-serif;
        text-shadow: 1px 1px 3px rgba(0,0,0,0.3);
        line-height: 1.2 !important;
    }
    .nav-bar p {
        margin: 2px 0 0 0 !important;
        padding: 0 !important;
        font-size: 13px !important;
        color: #E0E7FF;
        font-weight: 500;
        line-height: 1.2 !important;
    }
    .block-container {
        padding-top: 5rem !important; /* Push content down safely */
        padding-bottom: 0rem !important;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 15px;
        margin-bottom: 20px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 55px;
        white-space: pre-wrap;
        background-color: #F1F5F9;
        border-radius: 8px 8px 0 0;
        padding: 10px 25px;
        color: #475569;
        font-weight: 600;
        border: 1px solid #E2E8F0;
        border-bottom: none;
    }
    .stTabs [aria-selected="true"] {
        background-color: white !important;
        box-shadow: 0 -4px 10px rgba(0,0,0,0.05);
        color: #1E3A8A !important;
        border-top: 3px solid #1E3A8A;
    }
</style>

<div class="nav-bar">
    <h1>GST Refund Automation Engine (Rule 89(5))</h1>
    <p>Automate your Max Permissible Refund calculation under Section 54(3) of the CGST Act with strictly accurate Invoice-Level Categorization.</p>
</div>
""", unsafe_allow_html=True)

tab_guide, tab1, tab2, tab3, tab_s1a = st.tabs(["📖 User Guide", "📁 Data Upload", "⚙️ Refund Calculator", "📄 Official Statement 1", "Statement 1A Excel"])

with tab_guide:
    try:
        with open("USER_GUIDE.md", "r", encoding="utf-8") as f:
            st.markdown(f.read())
    except FileNotFoundError:
        st.info("User Guide not found. Please ensure USER_GUIDE.md is in the same directory.")

with tab1:
    st.header("Upload Tax Documents")
    
    gstr2b_path = st.file_uploader("Upload GSTR-2B (Excel)", type=["xlsx"])
    gstr1_path = st.file_uploader("Upload GSTR-1 (JSON)", type=["json"], accept_multiple_files=True)
    
    if not gstr2b_path or not gstr1_path:
        st.info("Please provide both GSTR-2B and GSTR-1 files to proceed to the next tabs.")
        st.stop()
    else:
        st.success("✅ Files successfully loaded! You can now proceed to the 'Refund Calculator' tab.")

@st.cache_data
def load_gstr2b(file_obj):
    try:
        # Read all sheets at once to avoid exhausting the file pointer of UploadedFile objects
        sheets_dict = pd.read_excel(file_obj, sheet_name=None, header=None)
        
        def process_sheet(df):
            if df is None or df.empty: return pd.DataFrame()
            h_idx = next((i for i, r in df.iterrows() if any('GSTIN' in str(v) for v in r.values)), -1)
            if h_idx != -1:
                cols = [f"{str(c1).strip()} {str(c2).strip()}".strip() for c1, c2 in zip(df.iloc[h_idx].fillna(''), df.iloc[h_idx+1].fillna(''))]
                res = df.iloc[h_idx+2:].copy()
                res.columns = cols
                return res
            return pd.DataFrame()
            
        b2b = process_sheet(sheets_dict.get('B2B', pd.DataFrame()))
        cdnr = process_sheet(sheets_dict.get('B2B-CDNR', pd.DataFrame()))
        
        gstin = "Unknown"
        legal_name = "Unknown"
        if 'Read me' in sheets_dict:
            readme = sheets_dict['Read me']
            for _, row in readme.iterrows():
                row_vals = [str(x).strip() for x in row.values if pd.notna(x)]
                for i, val in enumerate(row_vals):
                    if val == 'GSTIN' and i + 1 < len(row_vals):
                        gstin = row_vals[i+1]
                    if val == 'Legal Name' and i + 1 < len(row_vals):
                        legal_name = row_vals[i+1]
                        
        return b2b, cdnr, gstin, legal_name
    except Exception as e:
        st.error(f"Error reading GSTR-2B: {e}")
        return None, None, "Unknown", "Unknown"

b2b_df, cdnr_df, user_gstin, user_legal_name = load_gstr2b(gstr2b_path)

@st.cache_data
def process_gstr1(file_objs):
    sales = []
    if not isinstance(file_objs, list):
        file_objs = [file_objs]
        
    for file_obj in file_objs:
        try:
            data = json.load(open(file_obj, "r", encoding="utf-8")) if isinstance(file_obj, str) else json.load(file_obj)
            
            if 'b2b' in data:
                for c in data['b2b']:
                    for inv in c.get('inv', []):
                        for itm in inv.get('itms', []):
                            det = itm.get('itm_det', {})
                            sales.append({'Type': 'B2B', 'Rate': det.get('rt', 0), 'Taxable Value': det.get('txval', 0), 'Tax': det.get('iamt', 0) + det.get('camt', 0) + det.get('samt', 0), 'IGST': det.get('iamt', 0), 'CGST': det.get('camt', 0), 'SGST': det.get('samt', 0)})
            if 'b2cs' in data:
                for c in data['b2cs']:
                    sales.append({'Type': 'B2CS', 'Rate': c.get('rt', 0), 'Taxable Value': c.get('txval', 0), 'Tax': c.get('iamt', 0) + c.get('camt', 0) + c.get('samt', 0), 'IGST': c.get('iamt', 0), 'CGST': c.get('camt', 0), 'SGST': c.get('samt', 0)})
            if 'b2cl' in data:
                for c in data['b2cl']:
                    for inv in c.get('inv', []):
                        for itm in inv.get('itms', []):
                            det = itm.get('itm_det', {})
                            sales.append({'Type': 'B2CL', 'Rate': det.get('rt', 0), 'Taxable Value': det.get('txval', 0), 'Tax': det.get('iamt', 0) + det.get('camt', 0) + det.get('samt', 0), 'IGST': det.get('iamt', 0), 'CGST': det.get('camt', 0), 'SGST': det.get('samt', 0)})
            if 'cdnr' in data:
                for c in data['cdnr']:
                    for nt in c.get('nt', []):
                        nt_type = nt.get('ntty')
                        mult = -1 if nt_type == 'C' else 1
                        for itm in nt.get('itms', []):
                            det = itm.get('itm_det', {})
                            sales.append({'Type': f'CDNR ({nt_type})', 'Rate': det.get('rt', 0), 'Taxable Value': det.get('txval', 0) * mult, 'Tax': (det.get('iamt', 0) + det.get('camt', 0) + det.get('samt', 0)) * mult, 'IGST': det.get('iamt', 0) * mult, 'CGST': det.get('camt', 0) * mult, 'SGST': det.get('samt', 0) * mult})
        except Exception as e:
            st.error(f"Error reading one of the GSTR-1 JSON files: {e}")
            
    try:
        df = pd.DataFrame(sales)
        return df
    except Exception as e:
        st.error(f"Error combining GSTR-1 Data: {e}")
        return pd.DataFrame()

sales_df = process_gstr1(gstr1_path)

val_col = next((col for col in b2b_df.columns if 'Invoice Value' in str(col) or 'Taxable Value' in str(col)), None) if b2b_df is not None else None
igst_col = next((col for col in b2b_df.columns if 'Integrated Tax' in str(col)), None) if b2b_df is not None else None
cgst_col = next((col for col in b2b_df.columns if 'Central Tax' in str(col)), None) if b2b_df is not None else None
sgst_col = next((col for col in b2b_df.columns if 'State/UT Tax' in str(col)), None) if b2b_df is not None else None
rate_col_b2b = 'Calculated_Rate'
txval_col = next((col for col in b2b_df.columns if 'Taxable' in str(col) and 'Value' in str(col)), val_col) if b2b_df is not None else None


with tab2:
    st.header("Phase 1: Inward Supply Processing (GSTR-2B)")

    if b2b_df is not None and not b2b_df.empty:
        supplier_col = next((col for col in b2b_df.columns if 'Trade/Legal name' in str(col) or 'Supplier Name' in str(col) or 'Trade Name' in str(col)), None)
        gstin_col = next((col for col in b2b_df.columns if 'GSTIN' in str(col)), None)
        invoice_col = next((col for col in b2b_df.columns if 'Invoice number' in str(col)), None)
        
        if supplier_col and gstin_col and invoice_col:
            st.subheader("Categorize ITC (Invoice-Level)")
            st.write("Assign the Type of ITC for **each invoice**. Capital Goods are excluded from the Net ITC calculation.")
            
            unique_invoices = b2b_df[[gstin_col, supplier_col, invoice_col, val_col]].copy()
            unique_invoices['ID'] = unique_invoices[gstin_col].astype(str) + "_" + unique_invoices[invoice_col].astype(str)
            
            type_col_exists = next((col for col in b2b_df.columns if 'Type of ITC' in str(col)), None)
            if type_col_exists:
                unique_invoices['Type of ITC'] = b2b_df[type_col_exists].fillna('Input Goods')
                st.success("✅ 'Type of ITC' column detected in your Excel file! Pre-filled invoice categories automatically.")
                b2b_df.drop(columns=[type_col_exists], inplace=True)
            else:
                unique_invoices['Type of ITC'] = 'Input Goods'
                
            edited_invoices = st.data_editor(
                unique_invoices.drop(columns=['ID']), 
                column_config={
                    "Type of ITC": st.column_config.SelectboxColumn(
                        "Type of ITC",
                        help="Select the category of ITC for this invoice",
                        options=["Input Goods", "Input Services", "Capital Goods"],
                        required=True,
                    )
                },
                hide_index=True,
                use_container_width=True
            )
            
            unique_invoices['Type of ITC'] = edited_invoices['Type of ITC']
            b2b_df['ID'] = b2b_df[gstin_col].astype(str) + "_" + b2b_df[invoice_col].astype(str)
            b2b_df = b2b_df.merge(unique_invoices[['ID', 'Type of ITC']], on='ID', how='left')
            
            b2b_df['Total Tax'] = b2b_df[[igst_col, cgst_col, sgst_col]].fillna(0).sum(axis=1)
            
            def calculate_gst_rate(tax, txval):
                try:
                    if float(txval) <= 0: return 0.0
                    implied = (float(tax) / float(txval)) * 100
                    return float(min([0, 5, 12, 18, 28], key=lambda x: abs(x - implied)))
                except:
                    return 0.0
                    
            b2b_df['Calculated_Rate'] = b2b_df.apply(lambda r: calculate_gst_rate(r['Total Tax'], r[txval_col]), axis=1)
            
            if not cdnr_df.empty:
                cdnr_gstin_col = next((col for col in cdnr_df.columns if 'GSTIN' in str(col)), None)
                cdnr_supp_col = next((col for col in cdnr_df.columns if 'Trade/Legal name' in str(col) or 'Supplier' in str(col)), None)
                cdnr_note_col = next((col for col in cdnr_df.columns if 'Note number' in str(col)), None)
                cdnr_val_col = next((col for col in cdnr_df.columns if 'Note Value' in str(col)), None)
                
                if cdnr_gstin_col and cdnr_note_col:
                    st.subheader("Categorize Credit/Debit Notes (CDNR)")
                    unique_notes = cdnr_df[[cdnr_gstin_col, cdnr_supp_col, cdnr_note_col, cdnr_val_col]].copy()
                    unique_notes['ID'] = unique_notes[cdnr_gstin_col].astype(str) + "_" + unique_notes[cdnr_note_col].astype(str)
                    
                    type_col_cdnr = next((col for col in cdnr_df.columns if 'Type of ITC' in str(col)), None)
                    if type_col_cdnr:
                        unique_notes['Type of ITC'] = cdnr_df[type_col_cdnr].fillna('Input Goods')
                        cdnr_df.drop(columns=[type_col_cdnr], inplace=True)
                    else:
                        unique_notes['Type of ITC'] = 'Input Goods'
                    
                    edited_notes = st.data_editor(
                        unique_notes.drop(columns=['ID']),
                        column_config={"Type of ITC": st.column_config.SelectboxColumn("Type of ITC", options=["Input Goods", "Input Services", "Capital Goods"], required=True)},
                        hide_index=True, use_container_width=True
                    )
                    
                    unique_notes['Type of ITC'] = edited_notes['Type of ITC']
                    cdnr_df['ID'] = cdnr_df[cdnr_gstin_col].astype(str) + "_" + cdnr_df[cdnr_note_col].astype(str)
                    cdnr_df = cdnr_df.merge(unique_notes[['ID', 'Type of ITC']], on='ID', how='left')
                    
                    cdnr_igst = next((col for col in cdnr_df.columns if 'Integrated Tax' in str(col)), None)
                    cdnr_cgst = next((col for col in cdnr_df.columns if 'Central Tax' in str(col)), None)
                    cdnr_sgst = next((col for col in cdnr_df.columns if 'State/UT Tax' in str(col)), None)
                    cdnr_cess = next((col for col in cdnr_df.columns if 'Cess' in str(col)), None)
                    
                    cdnr_df['Total Tax CDNR'] = cdnr_df[[cdnr_igst, cdnr_cgst, cdnr_sgst]].fillna(0).sum(axis=1)
                    
                    note_type_col = next((col for col in cdnr_df.columns if 'Document Type' in str(col) or 'Note Type' in str(col) or 'Note type' in str(col)), None)
                    
                    def apply_mult(row, col):
                        val = row.get(col, 0)
                        if isinstance(val, pd.Series):
                            val = val.iloc[0]
                        if pd.isna(val): return 0
                        
                        nt = row.get(note_type_col, '')
                        if isinstance(nt, pd.Series):
                            nt = nt.iloc[0]
                            
                        return -val if 'Credit' in str(nt) else val

                    if note_type_col:
                        cdnr_df['Adjustment'] = cdnr_df.apply(lambda r: apply_mult(r, 'Total Tax CDNR'), axis=1)
                        cdnr_df['Adj_IGST'] = cdnr_df.apply(lambda r: apply_mult(r, cdnr_igst) if cdnr_igst else 0, axis=1)
                        cdnr_df['Adj_CGST'] = cdnr_df.apply(lambda r: apply_mult(r, cdnr_cgst) if cdnr_cgst else 0, axis=1)
                        cdnr_df['Adj_SGST'] = cdnr_df.apply(lambda r: apply_mult(r, cdnr_sgst) if cdnr_sgst else 0, axis=1)
                        cdnr_df['Adj_CESS'] = cdnr_df.apply(lambda r: apply_mult(r, cdnr_cess) if cdnr_cess else 0, axis=1)
                    else:
                        cdnr_df['Adjustment'] = -cdnr_df['Total Tax CDNR']
                        cdnr_df['Adj_IGST'] = -cdnr_df[cdnr_igst].fillna(0) if cdnr_igst else 0
                        cdnr_df['Adj_CGST'] = -cdnr_df[cdnr_cgst].fillna(0) if cdnr_cgst else 0
                        cdnr_df['Adj_SGST'] = -cdnr_df[cdnr_sgst].fillna(0) if cdnr_sgst else 0
                        cdnr_df['Adj_CESS'] = -cdnr_df[cdnr_cess].fillna(0) if cdnr_cess else 0
            else:
                cdnr_df = pd.DataFrame(columns=['Type of ITC', 'Adjustment', 'Adj_IGST', 'Adj_CGST', 'Adj_SGST', 'Adj_CESS'])
                
            b2b_goods = b2b_df[b2b_df['Type of ITC'] == 'Input Goods']
            cess_col = next((col for col in b2b_df.columns if 'Cess' in str(col)), None)
            
            igst_goods = b2b_goods[igst_col].fillna(0).sum() if igst_col else 0
            cgst_goods = b2b_goods[cgst_col].fillna(0).sum() if cgst_col else 0
            sgst_goods = b2b_goods[sgst_col].fillna(0).sum() if sgst_col else 0
            cess_goods = b2b_goods[cess_col].fillna(0).sum() if cess_col else 0
            
            cdnr_goods = cdnr_df[cdnr_df['Type of ITC'] == 'Input Goods'] if not cdnr_df.empty else pd.DataFrame()
            if not cdnr_goods.empty:
                igst_goods += cdnr_goods['Adj_IGST'].sum()
                cgst_goods += cdnr_goods['Adj_CGST'].sum()
                sgst_goods += cdnr_goods['Adj_SGST'].sum()
                cess_goods += cdnr_goods['Adj_CESS'].sum()
                
            input_goods_tax = b2b_goods['Total Tax'].sum()
            input_services_tax = b2b_df[b2b_df['Type of ITC'] == 'Input Services']['Total Tax'].sum()
            
            input_goods_adj = cdnr_goods['Adjustment'].sum() if not cdnr_goods.empty else 0
            input_services_adj = cdnr_df[cdnr_df['Type of ITC'] == 'Input Services']['Adjustment'].sum() if not cdnr_df.empty else 0
            
            net_itc_goods = input_goods_tax + input_goods_adj
            net_itc_services = input_services_tax + input_services_adj
            total_itc = net_itc_goods + net_itc_services
            
            col1, col2, col3 = st.columns(3)
            col1.metric("Net ITC (Goods Only)", f"₹ {net_itc_goods:,.2f}")
            col2.metric("Total ITC (Services)", f"₹ {net_itc_services:,.2f}")
            col3.metric("Total Eligible ITC", f"₹ {total_itc:,.2f}")

    st.header("Phase 2: Outward Supply Processing (GSTR-1 JSON)")

    if not sales_df.empty:
        rate_summary = sales_df.groupby('Rate').agg({'Taxable Value': 'sum', 'Tax': 'sum'}).reset_index()
        st.subheader("Rate-Wise Outward Supplies")
        st.dataframe(rate_summary.style.format({'Taxable Value': '{:,.2f}', 'Tax': '{:,.2f}'}), use_container_width=True)
        
        st.subheader("Inverted Rated Supply (IRS) Identification")
        available_rates = sorted(rate_summary['Rate'].unique().tolist())
        predominant_input_rate = st.selectbox("Select your Predominant Input Goods GST Rate (%):", [0.0, 5.0, 18.0, 28.0], index=2)
        
        auto_irs_rates = [r for r in available_rates if r < predominant_input_rate and r > 0]
        irs_rates = st.multiselect("Confirm IRS Output Rates (Auto-selected based on Input Rate):", available_rates, default=auto_irs_rates)
        
        turnover_irs = rate_summary[rate_summary['Rate'].isin(irs_rates)]['Taxable Value'].sum()
        tax_payable_irs = rate_summary[rate_summary['Rate'].isin(irs_rates)]['Tax'].sum()
        adj_total_turnover = rate_summary['Taxable Value'].sum()
        
        col1, col2, col3 = st.columns(3)
        col1.metric("Turnover of IRS", f"₹ {turnover_irs:,.2f}")
        col2.metric("Adjusted Total Turnover", f"₹ {adj_total_turnover:,.2f}")
        col3.metric("Tax Payable on IRS", f"₹ {tax_payable_irs:,.2f}")

        st.header("Phase 3: Final Refund Computation (Rule 89(5))")
        
        if 'total_itc' in locals() and adj_total_turnover > 0 and total_itc > 0:
            term1 = (turnover_irs * net_itc_goods) / adj_total_turnover
            term2 = tax_payable_irs * (net_itc_goods / total_itc)
            final_refund = max(0, term1 - term2)
            
            st.success(f"### Maximum Permissible Refund: ₹ {final_refund:,.2f}")
            with st.expander("Show Calculation Breakdown"):
                st.code(f"""
Refund = [(Turnover of IRS × Net ITC) / Adjusted Total Turnover] - [Tax Payable on IRS × (Net ITC / Total ITC)]
Refund = [({turnover_irs:,.2f} × {net_itc_goods:,.2f}) / {adj_total_turnover:,.2f}] - [{tax_payable_irs:,.2f} × ({net_itc_goods:,.2f} / {total_itc:,.2f})]
Refund = [{term1:,.2f}] - [{term2:,.2f}]
Refund = {final_refund:,.2f}
                """)
        else:
            st.warning("Ensure Adjusted Total Turnover and Total ITC are greater than zero. Complete Phase 1 to calculate ITC.")

with tab3:
    if 'total_itc' in locals() and 'sales_df' in locals() and not sales_df.empty and adj_total_turnover > 0 and total_itc > 0:
        
        # --- Generate Rate-Wise Inward Supply HTML ---
        inward_rate_html = ""
        if 'Type of ITC' in b2b_df.columns:
            b2b_df[val_col] = pd.to_numeric(b2b_df[val_col], errors='coerce').fillna(0)
            
            for itc_type in ['Input Goods', 'Input Services', 'Capital Goods']:
                type_df = b2b_df[b2b_df['Type of ITC'] == itc_type]
                if not type_df.empty:
                    txval = type_df[val_col].sum()
                    cgst = type_df[cgst_col].fillna(0).sum() if cgst_col else 0
                    sgst = type_df[sgst_col].fillna(0).sum() if sgst_col else 0
                    igst = type_df[igst_col].fillna(0).sum() if igst_col else 0
                    tot = type_df['Total Tax'].sum()
                    
                    if txval > 0 or tot > 0:
                        inward_rate_html += f"""
        <tr>
            <td style="border: 1px solid black; padding: 5px; text-align: left;">{itc_type}</td>
            <td style="border: 1px solid black; padding: 5px;">{txval:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px;">{cgst:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px;">{sgst:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px;">{igst:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px;">{tot:,.2f}</td>
        </tr>"""

        # --- Generate Rate-Wise Outward Supply HTML ---
        outward_rate_html = ""
        outward_grp = sales_df.groupby('Rate').agg({
            'Taxable Value': 'sum', 'CGST': 'sum', 'SGST': 'sum', 'IGST': 'sum', 'Tax': 'sum'
        }).reset_index()
        
        for _, row in outward_grp.iterrows():
            rate_val = str(row['Rate']).replace('%', '')
            txval = row['Taxable Value']
            cgst = row['CGST']
            sgst = row['SGST']
            igst = row['IGST']
            tot = row['Tax']
            if txval > 0 or tot > 0:
                outward_rate_html += f"""
        <tr>
            <td style="border: 1px solid black; padding: 5px; text-align: left;">Goods @ {rate_val}%</td>
            <td style="border: 1px solid black; padding: 5px;">{txval:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px;">{cgst:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px;">{sgst:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px;">{igst:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px;">{tot:,.2f}</td>
        </tr>"""

        # Build pure HTML string for both UI and PDF
        outward_txval, outward_cgst, outward_sgst, outward_igst, outward_tot = outward_grp['Taxable Value'].sum(), outward_grp['CGST'].sum(), outward_grp['SGST'].sum(), outward_grp['IGST'].sum(), outward_grp['Tax'].sum()
        html_report = textwrap.dedent(f"""\
<div style="background-color: white; padding: 10px; color: black; font-family: sans-serif;">
    <h2 style="text-align: center; text-decoration: underline; font-family: Arial, sans-serif;">Computation of GST Refund</h2>
    
    <table style="width: 100%; font-size: 11px; border-collapse: collapse; text-align: center; border: 1px solid black; color: black; font-family: Arial, sans-serif;">
        <tr style="background-color: #f8f9fa; font-weight: bold;">
            <td style="border: 1px solid black; padding: 5px; text-align: left; width: 16%; white-space: nowrap;">Trade Name :</td>
            <td colspan="5" style="border: 1px solid black; padding: 5px; text-align: left; width: 84%;">{str(user_legal_name).upper()}</td>
        </tr>
        <tr style="background-color: #f8f9fa; font-weight: bold;">
            <td style="border: 1px solid black; padding: 5px; text-align: left; white-space: nowrap;">GST No:</td>
            <td colspan="5" style="border: 1px solid black; padding: 5px; text-align: left;">{user_gstin}</td>
        </tr>
        <tr style="background-color: #f8f9fa; font-weight: bold;">
            <th style="border: 1px solid black; padding: 5px; text-align: left; white-space: nowrap;">Particular</th>
            <th style="border: 1px solid black; padding: 5px; white-space: nowrap;">Taxable Value</th>
            <th style="border: 1px solid black; padding: 5px; white-space: nowrap;">CGST</th>
            <th style="border: 1px solid black; padding: 5px; white-space: nowrap;">SGST</th>
            <th style="border: 1px solid black; padding: 5px; white-space: nowrap;">IGST</th>
            <th style="border: 1px solid black; padding: 5px; white-space: nowrap;">Total Tax</th>
        </tr>
        <tr>
            <td colspan="6" style="border: 1px solid black; padding: 5px; text-align: left; font-weight: bold; background-color: #f0f0f0;"><u>Total Inward Supply (Purchase)</u></td>
        </tr>
{inward_rate_html}
        <tr>
            <td colspan="6" style="border: 1px solid black; padding: 5px; text-align: left; font-weight: bold; background-color: #f0f0f0;"><u>Outward Supply (Sales)</u></td>
        </tr>
{outward_rate_html}
        <tr style="font-weight: bold; background-color: #f8f9fa;">
            <td style="border: 1px solid black; padding: 5px; text-align: left;">Net outward supply</td>
            <td style="border: 1px solid black; padding: 5px; white-space: nowrap;">{outward_txval:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px; white-space: nowrap;">{outward_cgst:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px; white-space: nowrap;">{outward_sgst:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px; white-space: nowrap;">{outward_igst:,.2f}</td>
            <td style="border: 1px solid black; padding: 5px; white-space: nowrap;">{outward_tot:,.2f}</td>
        </tr>
    </table>

    <h3 style="color: #2c3e50; margin-top: 1.5em; font-family: Arial, sans-serif;">COMPUTATION OF REFUND TO BE CLAIMED (STATEMENT 1)</h3>
    <table style="width: 100%; font-size: 11px; border-collapse: collapse; text-align: center; border: 1px solid black; color: black; font-family: Arial, sans-serif;">
        <tr style="background-color: #f8f9fa; font-weight: bold;">
            <th style="border: 1px solid black; padding: 4px; border-right: none;">&nbsp;</th>
            <th style="border: 1px solid black; padding: 4px;">Turnover of inverted rated<br>supply of goods and<br>services (1) (₹)</th>
            <th style="border: 1px solid black; padding: 4px;">Tax payable on such<br>inverted rated supply of<br>goods and services * (Net<br>ITC / ITC availed on<br>inputs and input services)<br>(2) (₹)</th>
            <th style="border: 1px solid black; padding: 4px;">Adjusted total turnover<br>(3) (₹)</th>
            <th style="border: 1px solid black; padding: 4px;">Net input tax credit<br>(4) (₹)<br><span style="font-size: 8px; color: #555555; font-weight: normal; display: block; margin-top: 4px;">Edit the Net ITC to exclude, the<br>ITC availed on input services and<br>capital goods and the ITC of<br>refund claimed under Rule 89(4A)<br>and/ or (4B)</span></th>
            <th style="border: 1px solid black; padding: 4px;">Maximum<br>refund<br>amount to<br>be claimed<br>(5)<br>[(1×4÷3)-<br>(2)]<br>(₹)</th>
        </tr>
        <tr>
            <td style="border: 1px solid black; padding: 4px; text-align: left; white-space: nowrap;">Integrated Tax</td>
            <td rowspan="3" style="border: 1px solid black; padding: 4px; background-color: #f0f0f0; white-space: nowrap;">₹ {turnover_irs:,.2f}</td>
            <td rowspan="3" style="border: 1px solid black; padding: 4px; background-color: #f0f0f0; white-space: nowrap;">₹ {term2:,.2f}</td>
            <td rowspan="3" style="border: 1px solid black; padding: 4px; background-color: #f0f0f0; white-space: nowrap;">₹ {adj_total_turnover:,.2f}</td>
            <td rowspan="3" style="border: 1px solid black; padding: 4px; background-color: #f0f0f0; white-space: nowrap;">₹ {net_itc_goods:,.2f}</td>
            <td rowspan="3" style="border: 1px solid black; padding: 4px; white-space: nowrap;">{final_refund:,.2f}</td>
        </tr>
        <tr>
            <td style="border: 1px solid black; padding: 4px; text-align: left; white-space: nowrap;">Central Tax</td>
        </tr>
        <tr>
            <td style="border: 1px solid black; padding: 4px; text-align: left; white-space: nowrap;">State/UT Tax</td>
        </tr>
        <tr>
            <td style="border: 1px solid black; padding: 4px; text-align: left; white-space: nowrap;">CESS</td>
            <td style="border: 1px solid black; padding: 4px; background-color: #f0f0f0;">₹0.00</td>
            <td style="border: 1px solid black; padding: 4px; background-color: #f0f0f0;">₹0.00</td>
            <td style="border: 1px solid black; padding: 4px; background-color: #f0f0f0;">₹0.00</td>
            <td style="border: 1px solid black; padding: 4px; background-color: #f0f0f0;">₹0.00</td>
            <td style="border: 1px solid black; padding: 4px;">0.00</td>
        </tr>
        <tr>
            <td style="border: 1px solid black; padding: 4px; text-align: left; font-weight: bold; white-space: nowrap;">Total</td>
            <td style="border: 1px solid black; padding: 4px; white-space: nowrap;">{turnover_irs:,.2f}</td>
            <td style="border: 1px solid black; padding: 4px; white-space: nowrap;">{term2:,.2f}</td>
            <td style="border: 1px solid black; padding: 4px; white-space: nowrap;">{adj_total_turnover:,.2f}</td>
            <td style="border: 1px solid black; padding: 4px; white-space: nowrap;">{net_itc_goods:,.2f}</td>
            <td style="border: 1px solid black; padding: 4px; font-weight: bold; white-space: nowrap;">{final_refund:,.2f}</td>
        </tr>
    </table>

    <div style="margin-top: 1em;">
        <p style="color: #c0392b; font-weight: bold; margin-bottom: 2px;">New formula as per Rule 89(5) &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;(W.e.f. 5th July 2022)*</p>
        <p style="color: #2980b9; font-weight: bold; margin-top: 0px; margin-bottom: 5px;">Circular No.181/13/2022 dated 10.11.2022</p>
        
        <p style="font-weight: bold; font-size: 11px; margin-bottom: 5px; font-family: Arial, sans-serif;">Calculation of Proportionate Tax payable As per Formula</p>
        <table style="width: 100%; font-size: 11px; border-collapse: collapse; text-align: center; border: 1px solid black; color: black; font-family: Arial, sans-serif;">
            <tr style="background-color: #ffffff;">
                <th style="border: 1px solid black; padding: 5px; width: 20%;">Turnover of inverted rated<br>supply of goods</th>
                <th style="border: 1px solid black; padding: 5px; width: 20%;">Tax payable on such inverted<br>rated supply of goods</th>
                <th style="border: 1px solid black; padding: 5px; width: 20%;">Net input tax credit</th>
                <th style="border: 1px solid black; padding: 5px; width: 20%;">Net ITC of Input and Input Services</th>
                <th style="border: 1px solid black; padding: 5px; width: 20%;">Proportionate Tax payable</th>
            </tr>
            <tr style="background-color: #ffffff; font-weight: bold;">
                <td style="border: 1px solid black; padding: 5px;">1</td>
                <td style="border: 1px solid black; padding: 5px;">2</td>
                <td style="border: 1px solid black; padding: 5px;">3</td>
                <td style="border: 1px solid black; padding: 5px;">4</td>
                <td style="border: 1px solid black; padding: 5px;">5</td>
            </tr>
            <tr>
                <td style="border: 1px solid black; padding: 5px; background-color: #fff200; font-weight: bold;">{turnover_irs:,.2f}</td>
                <td style="border: 1px solid black; padding: 5px; background-color: #fff200; font-weight: bold;">{tax_payable_irs:,.2f}</td>
                <td style="border: 1px solid black; padding: 5px; background-color: #fff200; font-weight: bold;">{net_itc_goods:,.2f}</td>
                <td style="border: 1px solid black; padding: 5px; background-color: #fce4d6; font-weight: bold;">{total_itc:,.2f}</td>
                <td style="border: 1px solid black; padding: 5px; background-color: #fce4d6; font-weight: bold;">{term2:,.2f}</td>
            </tr>
        </table>
    </div>
</div>
""")

        # --- EXPORT BUTTONS ---
        colA, colB = st.columns(2)
        
        file_name_prefix = f"STATEMENT_1_{user_gstin}_{str(user_legal_name)[:15].replace(' ', '_')}".upper()
        
        # 1. HTML as XLS (Excel will open this flawlessly matching the UI)
        html_for_excel = f"<html xmlns:x=\"urn:schemas-microsoft-com:office:excel\"><head><meta charset=\"UTF-8\"><style>body {{ font-family: Arial, sans-serif; }}</style></head><body>{html_report}</body></html>"
        colA.download_button("📊 Download as Excel (Formatted)", data=html_for_excel.encode('utf-8'), file_name=f"{file_name_prefix}.xls", mime="application/vnd.ms-excel", use_container_width=True)

        # 2. Real PDF Generation
        pdf_buffer = io.BytesIO()
        # Replace the ₹ symbol with Rs. exclusively for the PDF generation to avoid xhtml2pdf font encoding errors
        safe_pdf_report = html_report.replace('₹', 'Rs. ')
        
        # Force a page break right before Statement 1 so it starts cleanly on page 2
        safe_pdf_report = safe_pdf_report.replace(
            '<h3 style="color: #2c3e50; margin-top: 1.5em; font-family: Arial, sans-serif;">COMPUTATION OF REFUND TO BE CLAIMED (STATEMENT 1)</h3>',
            '<pdf:nextpage />\n<h3 style="color: #2c3e50; margin-top: 0; font-family: Arial, sans-serif;">COMPUTATION OF REFUND TO BE CLAIMED (STATEMENT 1)</h3>'
        )
        
        # Re-apply strict padding limits to completely prevent height explosion in xhtml2pdf
        pdf_css = """
        <style>
            @page { size: a4 landscape; margin: 0.5cm 0.5cm; }
            body, table, p, h2, h3, span { font-size: 9.5pt !important; margin: 0 !important; }
            td, th { padding: 4px !important; }
            h2 { font-size: 13pt !important; margin-bottom: 4px !important; margin-top: 0 !important; }
            h3 { font-size: 11pt !important; margin-bottom: 4px !important; margin-top: 4px !important; }
            .stApp { background: white !important; }
        </style>
        """
        pdf_html = f"<html><head><meta charset=\"UTF-8\">{pdf_css}</head><body>{safe_pdf_report}</body></html>"
        pisa.CreatePDF(io.StringIO(pdf_html), dest=pdf_buffer)
        pdf_data = pdf_buffer.getvalue()
        
        colB.download_button("🖨️ Download Actual PDF File", data=pdf_data, file_name=f"{file_name_prefix}.pdf", mime="application/pdf", use_container_width=True)

        # Render the Report Visual
        st.markdown("<br>", unsafe_allow_html=True)
        import streamlit.components.v1 as components
        components.html(html_report, height=1000, scrolling=True)
# --- NEW FEATURE: STATEMENT 1A EXCEL FILLER ---

def extract_invoice_rows_for_filler(gstr1_data_list):
    """Deep extraction logic from GSTR-1 JSON for Statement 1A."""
    rows = []
    for data in gstr1_data_list:
        try:
            if 'b2b' in data:
                for company in data['b2b']:
                    for inv in company.get('inv', []):
                        txval = sum(itm.get('itm_det', {}).get('txval', 0) for itm in inv.get('itms', []))
                        iamt = sum(itm.get('itm_det', {}).get('iamt', 0) for itm in inv.get('itms', []))
                        camt = sum(itm.get('itm_det', {}).get('camt', 0) for itm in inv.get('itms', []))
                        samt = sum(itm.get('itm_det', {}).get('samt', 0) for itm in inv.get('itms', []))
                        rows.append({
                            'type': 'B2B', 'no': inv.get('inum', ''), 'dt': inv.get('idt', ''),
                            'txval': txval, 'iamt': iamt, 'camt': camt, 'samt': samt
                        })
            if 'b2cl' in data:
                for state in data['b2cl']:
                    for inv in state.get('inv', []):
                        txval = sum(itm.get('itm_det', {}).get('txval', 0) for itm in inv.get('itms', []))
                        iamt = sum(itm.get('itm_det', {}).get('iamt', 0) for itm in inv.get('itms', []))
                        rows.append({
                            'type': 'B2C-Large', 'no': inv.get('inum', ''), 'dt': inv.get('idt', ''),
                            'txval': txval, 'iamt': iamt, 'camt': 0.0, 'samt': 0.0
                        })
            if 'b2cs' in data:
                for rec in data['b2cs']:
                    rows.append({
                        'type': 'B2C-Small', 'no': '', 'dt': '',
                        'txval': rec.get('txval', 0), 'iamt': rec.get('iamt', 0), 
                        'camt': rec.get('camt', 0), 'samt': rec.get('samt', 0)
                    })
            if 'cdnr' in data:
                for company in data['cdnr']:
                    for nt in company.get('nt', []):
                        nt_type = nt.get('ntty', 'C')
                        mult = -1 if nt_type == 'C' else 1
                        txval = sum(itm.get('itm_det', {}).get('txval', 0) for itm in nt.get('itms', [])) * mult
                        iamt = sum(itm.get('itm_det', {}).get('iamt', 0) for itm in nt.get('itms', [])) * mult
                        camt = sum(itm.get('itm_det', {}).get('camt', 0) for itm in nt.get('itms', [])) * mult
                        samt = sum(itm.get('itm_det', {}).get('samt', 0) for itm in nt.get('itms', [])) * mult
                        rows.append({
                            'type': 'B2B', 'no': nt.get('nt_num', '') or nt.get('ntnum', ''),
                            'dt': nt.get('nt_dt', '') or nt.get('ntdt', ''),
                            'txval': txval, 'iamt': iamt, 'camt': camt, 'samt': samt
                        })
        except: continue
    return rows

def generate_s1a_excel_xlsm(b2b_df, gstr1_json_list, gstin, from_period, to_period):
    """
    ULTRA-STABLE EXCEL ENGINE:
    Uses openpyxl with keep_vba=True to fill the XLSM template.
    Preserves all macros, buttons, and dropdowns.
    """
    import openpyxl
    TEMPLATE_NAME = "GST_REFUND_S01A.xlsm"
    
    if not os.path.exists(TEMPLATE_NAME):
        return None, f"Template {TEMPLATE_NAME} not found in root.", None
        
    try:
        # Load XLSM with VBA preservation
        wb = openpyxl.load_workbook(TEMPLATE_NAME, keep_vba=True)
        if "RFD_STMT01A" not in wb.sheetnames:
            return None, "Sheet 'RFD_STMT01A' missing in template.", None
        
        ws = wb["RFD_STMT01A"]
        
        # 1. Headers
        ws["C4"] = gstin
        ws["C5"] = from_period
        ws["C6"] = to_period
        
        # 2. Fill Inward Data (Rows 11+)
        if b2b_df is not None and not b2b_df.empty:
            # Dynamically find column names to avoid index errors
            g_c = next((c for c in b2b_df.columns if 'GSTIN' in str(c)), None)
            i_c = next((c for c in b2b_df.columns if 'Invoice number' in str(c)), None)
            d_c = next((c for c in b2b_df.columns if 'Invoice date' in str(c)), None)
            t_c = next((c for c in b2b_df.columns if 'Taxable' in str(c)), None)
            it_c = next((c for c in b2b_df.columns if 'Integrated' in str(c)), None)
            ct_c = next((c for c in b2b_df.columns if 'Central' in str(c)), None)
            st_c = next((c for c in b2b_df.columns if 'State' in str(c)), None)

            for idx, row in b2b_df.iterrows():
                r = 11 + idx
                ws.cell(row=r, column=1, value=idx + 1)
                ws.cell(row=r, column=2, value="Inward Supply from Registered Person")
                ws.cell(row=r, column=3, value=str(row.get(g_c, '')))
                ws.cell(row=r, column=4, value="Invoice/Bill of Entry")
                ws.cell(row=r, column=5, value=str(row.get(i_c, '')))
                ws.cell(row=r, column=6, value=str(row.get(d_c, '')))
                ws.cell(row=r, column=7, value=float(row.get(t_c, 0)))
                ws.cell(row=r, column=8, value=float(row.get(it_c, 0)))
                ws.cell(row=r, column=9, value=float(row.get(ct_c, 0)))
                ws.cell(row=r, column=10, value=float(row.get(st_c, 0)))

        # 3. Fill Outward Data (Rows 11+)
        outward_rows = extract_invoice_rows_for_filler(gstr1_json_list)
        for idx, orow in enumerate(outward_rows):
            r = 11 + idx
            ws.cell(row=r, column=12, value=orow['type'])
            ws.cell(row=r, column=13, value="Invoice")
            ws.cell(row=r, column=14, value=str(orow['no']))
            ws.cell(row=r, column=15, value=str(orow['dt']))
            ws.cell(row=r, column=16, value=float(orow['txval']))
            ws.cell(row=r, column=17, value=float(orow['iamt']))
            ws.cell(row=r, column=18, value=float(orow['camt']))
            ws.cell(row=r, column=19, value=float(orow['samt']))

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue(), None, "xlsm"
        
    except Exception as e:
        return None, f"Excel Filling Error: {e}", None

with tab_s1a:
    st.header("📥 Statement 1A Automation Utility")
    st.write("Automatically populate the official **Statement 1A XLSM Offline Tool** using your uploaded data.")
    
    with st.expander("📝 Business Details for Statement 1A", expanded=True):
        c1, c2 = st.columns(2)
        gstin_input = c1.text_input("GSTIN", value=user_gstin if user_gstin != "Unknown" else "", placeholder="Enter GSTIN")
        legal_name_input = c2.text_input("Legal Name", value=user_legal_name if user_legal_name != "Unknown" else "", placeholder="Enter Legal Name")
        
        c3, c4 = st.columns(2)
        from_period = c3.text_input("From Period (mmyyyy)", placeholder="042024")
        to_period = c4.text_input("To Period (mmyyyy)", placeholder="032025")

    if st.button("🚀 Generate & Fill Statement 1A Excel", use_container_width=True, type="primary"):
        if not gstr1_path:
            st.error("Please upload GSTR-1 JSON files first.")
        else:
            with st.spinner("Surgically filling your Excel template..."):
                # Load JSON data from uploaded files
                gstr1_data_list = []
                for f in gstr1_path:
                    f.seek(0)
                    gstr1_data_list.append(json.load(f))
                
                excel_data, err, ext = generate_s1a_excel_xlsm(b2b_df, gstr1_data_list, gstin_input, from_period, to_period)
                
                if err:
                    st.error(f"Engine Error: {err}")
                else:
                    st.success("✅ Excel Statement 1A Generated Successfully!")
                    st.download_button(
                        label="💾 Download Filled Statement 1A (.xlsm)",
                        data=excel_data,
                        file_name=f"GST_REFUND_S1A_{gstin_input}.xlsm",
                        mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                        use_container_width=True
                    )
                    st.info("💡 Note: Open the file and click 'Validate' or 'Generate JSON' to use the government utility features.")
